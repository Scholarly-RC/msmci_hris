import json
import logging

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q
from django.db.utils import IntegrityError
from django.http import HttpResponse
from django.http.request import QueryDict
from django.shortcuts import redirect, render
from django.urls import reverse
from django_htmx.http import (
    HttpResponseClientRedirect,
    reswap,
    retarget,
    trigger_client_event,
)
from openpyxl import load_workbook
from render_block import render_block_to_string

from attendance.utils.date_utils import (
    get_current_local_date,
    get_date_object_from_date_str,
)
from core.actions import (
    process_add_app_log_entry,
    process_add_personal_file,
    process_change_profile_picture,
    process_delete_personal_file,
    process_get_or_create_intial_user_one_to_one_fields,
    process_update_user_and_user_details,
)
from core.decorators import hr_required
from core.models import Department, Notification, PersonalFile
from core.notification import mark_notification_read, user_has_unread_notification
from core.utils import (
    check_if_biometric_uid_exists,
    check_user_has_password,
    generate_username_from_employee_id,
    get_app_logs,
    get_civil_status_list,
    get_education_list,
    get_education_list_with_degrees_earned,
    get_gender_list,
    get_personal_file_categories,
    get_religion_list,
    get_role_list,
    get_user_personal_files,
    get_user_with_logs,
    get_users_sorted_by_department,
    profile_picture_validation,
)
from core.validations import password_validation, personal_file_validation
from hris.exceptions import PersonalFilesBlockNotFound
from hris.utils import create_global_alert_instance
from payroll.utils import get_rank_choices

logger = logging.getLogger(__name__)


@login_required(login_url="/login")
def main(request):
    context = {}
    return render(request, "core/main.html", context)


def user_login(request):
    if request.user.is_authenticated:
        return redirect(reverse("core:profile"))

    context = {}
    if request.method == "POST":
        email = request.POST["email"]
        password = request.POST["password"]
        # check_user_has_password(email=email)
        user = authenticate(request, username=email, password=password)
        response = HttpResponse()
        if user is not None:
            login(request, user)
            process_add_app_log_entry(request.user.id, "Logged in.")
            response = HttpResponseClientRedirect(reverse("core:main"))
            return response
        else:
            if not password:
                if not check_user_has_password(email=email):
                    context.update({"user_email": email})
                    response.content = render_block_to_string(
                        "core/components/set_password.html",
                        "set_password_form",
                        context,
                    )
                    response = retarget(response, "#login_card")
                    response = reswap(response, "outerHTML")
                    return response

            context.update(
                {
                    "login_error_message": "Username or password is incorrect. Please try again.",
                }
            )
            response.content = render_block_to_string(
                "core/login.html", "login_error_message", context
            )
            response = retarget(response, "#login_error_message")
            response = reswap(response, "outerHTML")
            return response

    return render(request, "core/login.html", context)


@login_required(login_url="/login")
def set_new_user_password(request):
    context = {}
    password = request.POST["password"]
    confirm_password = request.POST["confirm_password"]
    response = HttpResponse()

    errors = password_validation(password, confirm_password)

    if errors:
        context.update({"password_validation_errors": errors})
        response.content = render_block_to_string(
            "core/components/set_password.html", "set_error_message", context
        )
        response = retarget(response, "#error_list")
        response = reswap(response, "outerHTML")
        return response

    user_email = request.POST["email"]
    user = User.objects.get(email=user_email)
    user.set_password(password)
    user.save()

    context.update(
        {
            "password_successfully_set_message": "Your password has been set. Please login."
        }
    )
    process_add_app_log_entry(request.user.id, "Changed account password.")
    response.content = render_block_to_string("core/login.html", "login_card", context)
    response = retarget(response, "#login_card")
    response = reswap(response, "outerHTML")
    return response


def user_logout(request):
    process_add_app_log_entry(request.user.id, "Logged out.")
    logout(request)
    response = HttpResponse()
    response = HttpResponseClientRedirect(reverse("core:login"))
    return response


def user_register(request):
    if request.user.is_authenticated:
        return redirect(reverse("core:profile"))

    context = {}
    if request.htmx and request.POST:
        data = request.POST
        response = HttpResponse()
        email = data.get("email")
        employee_id = data.get("employee_id")
        password = data.get("password")
        confirm_password = data.get("confirm_password")
        errors = password_validation(password, confirm_password)
        if errors:
            context.update(
                {
                    "register_error_message": errors[0],
                }
            )
            response.content = render_block_to_string(
                "core/register.html", "register_error_message", context
            )
            response = retarget(response, "#register_error_message")
            response = reswap(response, "outerHTML")
            return response

        user, created = User.objects.get_or_create(
            email=email,
            defaults={"username": generate_username_from_employee_id(employee_id)},
        )
        if not created:
            context.update(
                {
                    "register_error_message": "The email address you entered is already registered. Please use a different email address or try logging in.",
                }
            )
            response.content = render_block_to_string(
                "core/register.html", "register_error_message", context
            )
            response = retarget(response, "#register_error_message")
            response = reswap(response, "outerHTML")
            return response

        user.set_password(password)
        user.save()

        user_details, _, _ = process_get_or_create_intial_user_one_to_one_fields(user)
        user_details[0].employee_number = employee_id
        user_details[0].save()

        context.update(
            {
                "account_successfully_created_message": "Account successfully created. Please login.",
            }
        )
        response.content = render_block_to_string(
            "core/register.html", "register_card", context
        )
        response = retarget(response, "#register_card")
        response = reswap(response, "outerHTML")
        return response

    return render(request, "core/register.html", context)


### USER PROFILE ###
@login_required(login_url="/login")
def user_profile(request):
    user = request.user
    departments = Department.objects.filter(is_active=True).order_by("name")
    education_list = get_education_list()
    civil_status_list = get_civil_status_list()
    religion_list = get_religion_list()
    roles = get_role_list()
    genders = get_gender_list()
    personal_files = get_user_personal_files(user=user)

    context = {
        "current_user": user,
        "department_list": departments,
        "civil_status_list": civil_status_list,
        "education_list": education_list,
        "religion_list": religion_list,
        "role_list": roles,
        "genders_list": genders,
        "personal_files": personal_files,
    }

    if (
        not request.htmx
        and user.userdetails.education in get_education_list_with_degrees_earned()
    ):
        context.update({"show_degrees_earned_section": True})

    if request.method == "POST":
        data = request.POST
        response = HttpResponse()
        if "toggle_degrees_earned" in data:
            if data.get("education") in get_education_list_with_degrees_earned():
                context.update({"show_degrees_earned_section": True})
            response.content = render_block_to_string(
                "core/user_profile.html", "degrees_earned_section", context
            )
            response = retarget(response, "#degrees_earned_section")
        else:
            updated_user, full_user_details = process_update_user_and_user_details(
                user_instance=user, querydict=data
            )

            if user.userdetails.education in get_education_list_with_degrees_earned():
                context.update({"show_degrees_earned_section": True})

            process_add_app_log_entry(
                request.user.id,
                f"Updated account profile information. Updated details: {full_user_details}.",
            )

            response = create_global_alert_instance(
                response, "Profile successfully updated.", "SUCCESS"
            )
            response.content = render_block_to_string(
                "core/user_profile.html", "profile_information_section", context
            )

            response = retarget(response, "#profile_information_section")
        response = reswap(response, "outerHTML")
        return response

    return render(request, "core/user_profile.html", context)


@login_required(login_url="/login")
def change_user_password(request):
    context = {}
    user = request.user
    if request.method == "POST":
        try:
            current_password = request.POST.get("current_password", "").strip()
            new_password = request.POST["new_password"].strip()
            confirm_password = request.POST["confirm_password"].strip()

            response = HttpResponse()
            response = retarget(response, "#password_information_section")
            response = reswap(response, "outerHTML")

            modify_selected_user = "modify_selected_user" in request.POST
            if modify_selected_user:
                user = User.objects.get(id=request.POST["modify_selected_user"])
                context["selected_user"] = user
                response.content = render_block_to_string(
                    "core/modify_user_profile.html",
                    "password_information_section",
                    context,
                )
            else:
                response.content = render_block_to_string(
                    "core/user_profile.html", "password_information_section", context
                )

            if new_password != confirm_password:
                response = create_global_alert_instance(
                    response, "Passwords do not match.", "WARNING"
                )
                return response

            if modify_selected_user:
                correct_password = True
            else:
                correct_password = user.check_password(current_password)

            if correct_password:
                if current_password == new_password and not modify_selected_user:
                    response = create_global_alert_instance(
                        response,
                        "New password cannot be the same as your old password.",
                        "WARNING",
                    )
                    return response
                else:
                    user.set_password(confirm_password)
                    user.save()
                    process_add_app_log_entry(
                        request.user.id,
                        (
                            "Changed account password."
                            if not modify_selected_user
                            else f"Changed user #{user.id} password."
                        ),
                    )
                    response = create_global_alert_instance(
                        response,
                        (
                            "Your password have been successfully changed. You will now be logged out of the site. Please refresh the page and log in again to continue."
                            if not modify_selected_user
                            else "Selected user's password have been successfully changed."
                        ),
                        "SUCCESS",
                    )
                    return response
            else:
                response = create_global_alert_instance(
                    response, "Current password is incorrect.", "WARNING"
                )
                return response
        except Exception as error:
            logger.error(
                "An error occurred while changing the user password", exc_info=True
            )
            response = create_global_alert_instance(
                response,
                "An error occurred while processing your request. Please try again.",
                "DANGER",
            )
            return response


@login_required(login_url="/login")
def upload_user_profile_picture(request):
    context = {}
    if request.FILES:
        profile_picture = request.FILES.get("profile_picture")
        error = profile_picture_validation(profile_picture)

        response = HttpResponse()
        response = retarget(response, "#profile_picture_section")
        response = reswap(response, "outerHTML")

        if error:
            response = create_global_alert_instance(response, error, "WARNING")
            response.content = render_block_to_string(
                "core/user_profile.html", "profile_picture_section", context
            )

        if error is None:
            user_details = request.user.userdetails
            user_details = process_change_profile_picture(profile_picture, user_details)
            process_add_app_log_entry(
                request.user.id,
                "Uploaded account profile picture.",
            )
            response = create_global_alert_instance(
                response, "Profile picture successfully uploaded.", "SUCCESS"
            )

            context.update(
                {
                    "new_profile_picture": user_details.profile_picture.url,
                }
            )
            response.content = render_block_to_string(
                "core/user_profile.html", "profile_picture_section", context
            )

    return response


@login_required(login_url="/login")
def add_personal_files(request):
    context = {}
    user = request.user
    if request.htmx:
        response = HttpResponse()
        if request.method == "GET":
            file_categories = get_personal_file_categories()
            context["file_categories"] = file_categories
            response.content = render_block_to_string(
                "core/user_profile.html", "add_file_modal_content", context
            )
            response = trigger_client_event(
                response, "initializeAddFileModal", after="swap"
            )
            response = retarget(response, "#add_file_modal_content")
            response = reswap(response, "outerHTML")
            return response

        if request.method == "POST":
            data = request.POST
            file_data = request.FILES
            try:
                errors = personal_file_validation(file_data=file_data)
                if errors:
                    response = create_global_alert_instance(
                        response,
                        f"{errors[0]}",
                        "WARNING",
                    )
                    response = reswap(response, "none")
                    return response

                new_file = process_add_personal_file(
                    user=user, payload=data, file_data=file_data
                )
                process_add_app_log_entry(
                    request.user.id, f"Added a personal file ({new_file.name})."
                )
                response = create_global_alert_instance(
                    response,
                    "Personal file has been successfully added.",
                    "SUCCESS",
                )
                response = trigger_client_event(
                    response, "closeAddFileModal", after="swap"
                )
                response = trigger_client_event(
                    response,
                    "reloadPersonalFilesSection",
                    {"category": new_file.category},
                    after="swap",
                )
                response = reswap(response, "none")
                return response
            except Exception as error:
                response = create_global_alert_instance(
                    response,
                    f"An error occurred while adding the personal file. Error details: {error}.",
                    "DANGER",
                )
                response = reswap(response, "none")
                return response


@login_required(login_url="/login")
def change_selected_category(request):
    context = {}
    if request.htmx and request.method == "POST":
        response = HttpResponse()
        data = request.POST
        selected_category = data.get("selected_category")
        if selected_category in [
            PersonalFile.PersonalFileCategory.DIPLOMA.value,
            PersonalFile.PersonalFileCategory.TRANSCRIPT_OF_RECORDS.value,
        ]:
            context["academic_degrees"] = get_education_list()

        if selected_category in [
            PersonalFile.PersonalFileCategory.CERTIFICATES.value,
            PersonalFile.PersonalFileCategory.MEDICAL_RECORDS.value,
        ]:
            context["show_year"] = True

        response.content = render_block_to_string(
            "core/user_profile.html", "additional_personal_file_details", context
        )

        response = retarget(response, "#additional_personal_file_details")
        response = reswap(response, "outerHTML")
        return response


@login_required(login_url="/login")
def delete_selected_personal_file(request):
    context = {}
    if request.htmx:
        response = HttpResponse()
        if request.method == "POST":
            data = request.POST
            file_id = data.get("file")
            file = PersonalFile.objects.get(id=file_id)
            context["file_to_delete"] = file

            response.content = render_block_to_string(
                "core/user_profile.html",
                "personal_file_delete_confirmation_modal_content",
                context,
            )

            response = trigger_client_event(
                response, "initializePersonalFileDeleteConfigmationModal", after="swap"
            )

            response = retarget(
                response, "#personal_file_delete_confirmation_modal_content"
            )
            response = reswap(response, "outerHTML")
            return response

        if request.method == "DELETE":
            data = QueryDict(request.body)
            try:
                category, file_name = process_delete_personal_file(payload=data)
                process_add_app_log_entry(
                    request.user.id, f"Deleted a personal file ({file_name})."
                )
                response = create_global_alert_instance(
                    response, "Selected Personal File successfully deleted.", "SUCCESS"
                )
                response = trigger_client_event(
                    response, "closePersonalFileDeleteConfigmationModal", after="swap"
                )
                response = trigger_client_event(
                    response,
                    "reloadPersonalFilesSection",
                    {"category": category},
                    after="swap",
                )
                return response
            except Exception as error:
                response = create_global_alert_instance(
                    response,
                    f"An error occurred while deleting the personal file. Error details: {error}.",
                    "DANGER",
                )
                response = reswap(response, "none")
                return response


@login_required(login_url="/login")
def reload_personal_files_section(request):
    context = {}
    user = request.user
    if request.htmx and request.method == "POST":
        response = HttpResponse()
        data = request.POST
        details = json.loads(data.get("details", {}))
        personal_files = get_user_personal_files(user=user)
        category = details.get("category")
        context["personal_files"] = personal_files

        if category == PersonalFile.PersonalFileCategory.DIPLOMA:
            block = "diploma_content"
        elif category == PersonalFile.PersonalFileCategory.TRANSCRIPT_OF_RECORDS:
            block = "tor_content"
        elif category == PersonalFile.PersonalFileCategory.ELIGIBILITY:
            block = "eligibility_content"
        elif category == PersonalFile.PersonalFileCategory.CERTIFICATES:
            block = "certificates_content"
        elif category == PersonalFile.PersonalFileCategory.SEMINARS_TRAININGS:
            block = "seminars_training_content"
        elif category == PersonalFile.PersonalFileCategory.MEDICAL_RECORDS:
            block = "medical_records_content"
        elif category == PersonalFile.PersonalFileCategory.OTHERS:
            block = "others_content"
        else:
            raise PersonalFilesBlockNotFound("Selected Personal File block is invalid.")

        response.content = render_block_to_string(
            "core/user_profile.html",
            block,
            context,
        )
        response = retarget(response, f"#{block}")
        response = reswap(response, "outerHTML")
        return response


@login_required(login_url="/login")
def preview_personal_file(request):
    context = {}
    if request.htmx and request.method == "POST":
        response = HttpResponse()
        data = request.POST
        file_id = data.get("file")
        file_to_preview = PersonalFile.objects.get(id=file_id)
        context["file_to_preview"] = file_to_preview

        response.content = render_block_to_string(
            "core/user_profile.html",
            "preview_personal_file_modal_content",
            context,
        )
        response = trigger_client_event(
            response, "initializePreviewPersonalFileModal", after="swap"
        )
        response = retarget(response, "#preview_personal_file_modal_content")
        response = reswap(response, "outerHTML")
        return response


### USER MANAGEMENT ###
@login_required(login_url="/login")
@hr_required("/")
def user_management(request):
    users = get_users_sorted_by_department()
    context = {"users": users}
    if request.htmx and request.POST:
        data = request.POST
        users_search = data.get("users_search", "")
        if users_search:
            user_filter = (
                Q(first_name__icontains=users_search)
                | Q(last_name__icontains=users_search)
                | Q(email__icontains=users_search)
            )
            users = users.filter(user_filter)
            context.update({"users": users})
        response = HttpResponse()
        response.content = render_block_to_string(
            "core/user_management.html",
            "user_management_table_content_container",
            context,
        )
        response = reswap(response, "outerHTML")
        response = retarget(response, "#user_management_table_content_container")
        return response

    return render(request, "core/user_management.html", context)


@login_required(login_url="/login")
def add_new_user(request):
    context = {}
    data = request.POST
    first_name = data["first_name"].strip()
    middle_name = data["middle_name"].strip()
    last_name = data["last_name"].strip()
    email = data["email"].strip().lower()
    employee_id = data["employee_id"].strip()

    if request.method == "POST":
        response = HttpResponse()
        try:
            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name,
                    "username": f"{first_name}{middle_name}{last_name}_temp",
                },
            )
            if created:
                user_details, _, _ = (
                    process_get_or_create_intial_user_one_to_one_fields(user)
                )
                user_details[0].middle_name = middle_name
                user_details[0].employee_number = employee_id
                user_details[0].save()
                user.username = generate_username_from_employee_id(employee_id)
                user.save()
                process_add_app_log_entry(
                    request.user.id, f"Added a new user (#{user.id})."
                )
                response = HttpResponseClientRedirect(reverse("core:user_management"))
            else:
                response = create_global_alert_instance(
                    response, "Email already exists.", "INFO"
                )
                response = reswap(response, "none")
        except IntegrityError as error:
            response = create_global_alert_instance(
                response, "User credentials already exists.", "INFO"
            )
            response = reswap(response, "none")
    return response


@login_required(login_url="/login")
def toggle_user_status(request, pk):
    user = User.objects.get(id=pk)
    user.is_active = not user.is_active
    user.save()

    process_add_app_log_entry(
        request.user.id,
        f"{'Disabled' if not user.is_active else 'Enabled'} user #{user.id}.",
    )

    context = {"user": user}
    response = HttpResponse()
    response.content = render_block_to_string(
        "core/user_management.html", "user_table_record", context
    )
    response = reswap(response, "outerHTML")
    return response


@login_required(login_url="/login")
def modify_user_details(request, pk):
    user = User.objects.get(id=pk)
    departments = Department.objects.filter(is_active=True).order_by("name")
    user_department = user.userdetails.department
    civil_status_list = get_civil_status_list()
    education_list = get_education_list()
    religion_list = get_religion_list()
    roles = get_role_list()
    genders = get_gender_list()
    personal_files = get_user_personal_files(user=user)

    context = {
        "selected_user": user,
        "department_list": departments,
        "civil_status_list": civil_status_list,
        "education_list": education_list,
        "religion_list": religion_list,
        "role_list": roles,
        "genders_list": genders,
        "personal_files": personal_files,
    }

    if user_department:
        rank_list = get_rank_choices(user_department.id)
        context.update({"rank_list": rank_list})

    if (
        not request.htmx
        and user.userdetails.education in get_education_list_with_degrees_earned()
    ):
        context.update({"show_degrees_earned_section": True})

    if request.POST:
        data = request.POST
        response = HttpResponse()
        if "toggle_degrees_earned" in data:
            if data.get("education") in get_education_list_with_degrees_earned():
                context.update({"show_degrees_earned_section": True})
            response.content = render_block_to_string(
                "core/modify_user_profile.html", "degrees_earned_section", context
            )
            response = retarget(response, "#degrees_earned_section")
        else:
            updated_user, full_user_details = process_update_user_and_user_details(
                user_instance=user, querydict=data
            )
            if user.userdetails.education in get_education_list_with_degrees_earned():
                context.update({"show_degrees_earned_section": True})

            process_add_app_log_entry(
                request.user.id,
                f"Updated the profile of user #{user.id}. New details: {full_user_details}.",
            )

            response = create_global_alert_instance(
                response, "Selected user successfully updated.", "SUCCESS"
            )
            context["selected_user"] = updated_user

            response.content = render_block_to_string(
                "core/modify_user_profile.html", "profile_information_section", context
            )
            response = retarget(response, "#profile_information_section")
        response = reswap(response, "outerHTML")
        return response

    return render(request, "core/modify_user_profile.html", context)


@login_required(login_url="/login")
def update_rank_selection(request):
    context = {}
    if request.htmx and request.method == "POST":
        data = request.POST
        response = HttpResponse()
        department_id = data.get("department")
        rank_list = get_rank_choices(department_id) if department_id else []
        context["rank_list"] = rank_list
        response.content = render_block_to_string(
            "core/modify_user_profile.html", "rank_selection", context
        )
        response = retarget(response, "#rank_selection")
        response = reswap(response, "outerHTML")
        return response


@login_required(login_url="/login")
def modify_user_biometric_details(request, pk):
    """
    Additional Info: This view is used by both User Profile and Change User Profile.
    """
    context = {}
    if request.htmx and request.POST:
        data = request.POST
        user = User.objects.get(id=pk)
        user_id_in_device = data.get("user_id_in_device", None)
        context["selected_user"] = user
        response = HttpResponse()
        if not check_if_biometric_uid_exists(current_user=user, uid=user_id_in_device):
            biometric_details = user.biometricdetail
            biometric_details.user_id_in_device = data.get("user_id_in_device", None)
            biometric_details.save()

            process_add_app_log_entry(
                request.user.id, f"Updated user #{user.id} biometric configuration."
            )

            response = create_global_alert_instance(
                response,
                "User biometric configuration successfully updated.",
                "SUCCESS",
            )
        else:
            response = create_global_alert_instance(
                response,
                "Provided Biometric UID is already in use by another user.",
                "WARNING",
            )
        if "user_profile" in data:
            context["current_user"] = user
            response.content = render_block_to_string(
                "core/user_profile.html", "biometric_configuration_section", context
            )
        else:
            response.content = render_block_to_string(
                "core/modify_user_profile.html",
                "biometric_configuration_section",
                context,
            )
        response = retarget(response, "#biometric_configuration_section")
        response = reswap(response, "outerHTML")
        return response


@login_required(login_url="/login")
def shift_modification_permision(request, pk):
    """
    Additional Info: This view is only used by Change User Profile.
    """
    context = {}
    if request.htmx and request.method == "POST":
        data = request.POST
        user = User.objects.get(id=pk)
        user_details = user.userdetails
        can_modify_shift = (
            data.get("can_modify_shift", False) and user_details.department is not None
        )
        user_details.can_modify_shift = can_modify_shift
        user_details.save()
        context["selected_user"] = user

        process_add_app_log_entry(
            request.user.id,
            f"{'Allowed' if can_modify_shift else 'Removed'} shift modification access from user ({user_details.get_user_fullname()}).",
        )

        response = HttpResponse()
        if user_details.department is not None:
            response = create_global_alert_instance(
                response,
                "User shift modification permission successfully updated.",
                "SUCCESS",
            )
        else:
            response = create_global_alert_instance(
                response,
                "A department must be assigned to the user before modifying shift modification permission.",
                "WARNING",
            )
        response.content = render_block_to_string(
            "core/modify_user_profile.html",
            "shift_modification_permision_section",
            context,
        )
        response = retarget(response, "#shift_modification_permision_section")
        response = reswap(response, "outerHTML")
        return response


@login_required(login_url="/login")
def bulk_add_new_users(request):
    context = {}
    if request.htmx and request.method == "POST" and request.FILES:
        excel_file = request.FILES["user_list"]

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active

            new_user_counter = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                email, employee_id = row[0], row[1]

                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        "username": generate_username_from_employee_id(employee_id)
                    },
                )
                if created:
                    new_user_counter += 1
                    process_get_or_create_intial_user_one_to_one_fields(user)

            if new_user_counter > 0:
                process_add_app_log_entry(
                    request.user.id, f"Bulk added {new_user_counter} new users."
                )
                messages.success(
                    request, message=f"{new_user_counter} users successfully added."
                )

            response = HttpResponseClientRedirect(reverse("core:user_management"))
            return response
        except Exception as error:
            logger.error("An error occurred while adding new users", exc_info=True)
            messages.error(
                request,
                "An error occurred while processing the file. Please check the format and try again.",
            )
            return redirect(reverse("core:user_management"))


### Notification Views ###
def notification_button_indicator(request):
    context = {}
    if request.htmx and request.method == "POST":
        user = request.user
        response = HttpResponse()
        has_unread_notification = user_has_unread_notification(user)
        context["has_unread_notification"] = has_unread_notification
        response.content = render_block_to_string(
            "navbar.html", "userNotificationButtonIndicator", context
        )
        response = retarget(response, "#userNotificationButtonIndicator")
        response = reswap(response, "outerHTML")
        return response


def retrieve_notifications(request):
    context = {}
    if request.htmx and request.method == "POST":
        response = HttpResponse()
        user = request.user
        notifications = user.notifications.order_by("-id")
        context["notifications"] = notifications
        response.content = render_block_to_string(
            "core/components/notifications.html", "userNotificationList", context
        )
        response = retarget(response, "#userNotificationList")
        response = reswap(response, "outerHTML")
        return response


def open_notification(request):
    context = {}
    if request.htmx and request.method == "POST":
        response = HttpResponse()
        try:
            data = request.POST
            notification_id = data.get("notification")
            notification = Notification.objects.get(id=notification_id)
            mark_notification_read(notification)
            response = HttpResponseClientRedirect(notification.url)
            return response
        except Exception as error:
            logger.error(f"Error accessing notification", exc_info=True)
            response = create_global_alert_instance(
                response,
                f"Something went wrong when accessing selected notification. Details: {error}.",
            )
            response = reswap(response, "none")
            return response


# App Log Views
@login_required(login_url="/login")
@hr_required("/")
def app_logs(request):
    context = {}
    selected_date = request.POST.get("app_log_date", None) or get_current_local_date()
    if type(selected_date) == str:
        selected_date = get_date_object_from_date_str(selected_date)

    user = request.POST.get("log_users", "0")

    logs = get_app_logs(selected_date, user)

    users = get_user_with_logs()

    context.update({"logs": logs, "users": users, "selected_date": selected_date})

    if request.htmx and request.method == "POST":
        response = HttpResponse()
        response.content = render_block_to_string(
            "core/app_logs.html", "app_log_table", context
        )
        response = retarget(response, "#app_log_table")
        response = reswap(response, "outerHTML")
        return response

    return render(request, "core/app_logs.html", context)


# App Shared View
def core_module_close_modals(request):
    if request.htmx and request.method == "POST":
        data = request.POST
        event_name = data.get("event_name")
        response = HttpResponse()
        response = trigger_client_event(response, event_name, after="swap")
        response = reswap(response, "none")
        return response
